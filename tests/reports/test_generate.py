import json
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from phiexplorer.extract.effectors import extract_effector_proteins
from phiexplorer.extract.phenotypes import extract_protein_phenotypes
from phiexplorer.reports.generate import (
    write_dataset_summary_report,
    write_effector_report,
    write_protein_phenotype_report,
)

FIXTURE_PATH = Path(__file__).parent.parent / "fixtures" / "sample_export.json"


@pytest.fixture
def export():
    with open(FIXTURE_PATH, encoding="utf-8") as f:
        return json.load(f)


def test_write_protein_phenotype_report(export, tmp_path):
    path = write_protein_phenotype_report(export, 90000, "Testus pathogenicus", output_dir=tmp_path)

    assert path.exists()
    assert path.name.startswith("testus_pathogenicus_protein_phenotypes_")
    assert path.suffix == ".xlsx"

    wb = openpyxl.load_workbook(path)
    ws = wb["Protein Phenotypes"]
    assert ws["A1"].value == "uniprot_id"
    assert ws["A2"].value == "Q00001"
    assert ws["A3"].value == "Q00002"

    df = extract_protein_phenotypes(export, 90000, "Testus pathogenicus")
    assert ws.max_row - 1 == len(df)
    header_row = [cell.value for cell in ws[1]]
    assert header_row == list(df.columns)


def test_write_effector_report(export, tmp_path):
    path = write_effector_report(export, 90000, "Testus pathogenicus", output_dir=tmp_path)

    assert path.exists()
    assert path.name.startswith("testus_pathogenicus_effectors_")
    assert path.suffix == ".xlsx"

    wb = openpyxl.load_workbook(path)
    ws = wb["Effectors"]
    assert ws["A1"].value == "uniprot_id"
    assert ws["A2"].value == "Q00001"

    df = extract_effector_proteins(export, 90000, "Testus pathogenicus")
    assert ws.max_row - 1 == len(df)
    header_row = [cell.value for cell in ws[1]]
    assert header_row == list(df.columns)


def test_write_dataset_summary_report(export, tmp_path):
    path = write_dataset_summary_report(export, output_dir=tmp_path)

    assert path.exists()
    assert path.name.startswith("dataset_summary_")
    assert path.suffix == ".csv"

    df = pd.read_csv(path, index_col=0)
    assert df.loc["Genes", "Count"] == 2
    assert df.loc["Pathogens", "Count"] == 1
    assert df.index.name == "Feature"


def test_write_protein_phenotype_report_default_output_dir(export, monkeypatch, tmp_path):
    monkeypatch.setattr("phiexplorer.reports.generate.paths.output_dir", lambda: tmp_path)

    path = write_protein_phenotype_report(export, 90000, "Testus pathogenicus")

    assert path.exists()
    assert path.parent == tmp_path


def test_write_dataset_summary_report_default_output_dir(export, monkeypatch, tmp_path):
    monkeypatch.setattr("phiexplorer.reports.generate.paths.output_dir", lambda: tmp_path)

    path = write_dataset_summary_report(export)

    assert path.exists()
    assert path.parent == tmp_path
