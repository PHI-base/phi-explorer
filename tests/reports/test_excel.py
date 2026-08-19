from pathlib import Path

import openpyxl
import pandas as pd

from phiexplorer.reports.excel import write_excel


def test_write_excel(tmp_path: Path):
    df = pd.DataFrame({"uniprot_id": ["Q00001", "Q00002"], "gene_name": ["geneA", "geneB"]})
    out_path = tmp_path / "test_output.xlsx"

    write_excel(df, out_path, sheet_name="Proteins")

    assert out_path.exists()
    wb = openpyxl.load_workbook(out_path)
    ws = wb["Proteins"]
    assert ws.freeze_panes == "A2"
    assert ws["A1"].value == "uniprot_id"
    assert ws["A2"].value == "Q00001"
